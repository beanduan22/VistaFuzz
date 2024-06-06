import random
import numpy as np
import cv2
import string
import struct
from datetime import datetime
from tool.oracle import *
import pandas as pd
import numpy as np
import random
import cv2
import string
import struct
from datetime import datetime
from tool.mutation_rules import *
import numpy as np
from openpyxl import load_workbook
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
np.set_printoptions(threshold=np.inf)
import os

class MutationHistory:
    def __init__(self, error_log_path):
        self.history = {}
        self.failed_combinations = set()
        self.mutation_strategies = ['random', 'incremental', 'bit_flip']
        self.error_log_path = error_log_path

    def add_record(self, api_name, param_type, success,mutated_arg, mutated_args, strategy):
        # 记录突变结果
        key = self._generate_key(mutated_arg)
        if not success:
            self.failed_combinations.add(key)
            #self.log_error(api_name, mutated_args, strategy)
        if param_type not in self.history:
            self.history[param_type] = {'successes': 0, 'failures': 0}
        if success:
            self.history[param_type]['successes'] += 1
        else:
            self.history[param_type]['failures'] += 1

    def suggest_mutation_strategy(self, param_type):
        # 基于历史记录调整突变策略
        data = self.history.get(param_type, {})
        failure_rate = data.get('failures', 0) / (data.get('successes', 0) + data.get('failures', 0) + 1)
        if failure_rate > 0.7:
            return 'bit_flip'
        elif failure_rate > 0.5:
            return random.choice(self.mutation_strategies)
        return 'incremental' if failure_rate > 0.2 else random.choice(self.mutation_strategies)

    def check_combination(self, mutated_arg):
        # 检查即将进行的突变是否已经导致了错误
        key = self._generate_key(mutated_arg)
        return key not in self.failed_combinations

    def log_error(self, api_name, mutated_arg,strategy,i):
        # 记录错误到日志文件
        error_record = pd.DataFrame([{
            'timestamp': datetime.now(),
            'api_name': api_name,
            'mutated_arg': str(mutated_arg),
            'strategy': strategy,

        }])
        bug_case_file = 'bug_case/'+api_name + '_' +str(i)+'.txt'
        if os.path.exists(self.error_log_path):
            # 读取原有数据
            with pd.ExcelWriter(self.error_log_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                # 如果使用ExcelWriter的append模式，确保`if_sheet_exists`参数为'overlay'或'replace'来避免错误
                # 这里我们读取原有数据是为了展示如何处理，实际上在追加模式下不需要先读取
                error_record.to_excel(writer, sheet_name='Sheet1', index=False, header=False,
                                      startrow=writer.sheets['Sheet1'].max_row)
        else:
            # 文件不存在，直接写入
            error_record.to_excel(self.error_log_path, index=False)
        with open(bug_case_file, 'a') as f:  # 打开文件以追加模式
            # 如果文件是空的，写入头部（列名）
            if f.tell() == 0:
                f.write(error_record.to_csv(sep='\t', index=False))
            else:
                # 文件不是空的，追加数据但不包括头部（列名）
                f.write(error_record.to_csv(sep='\t', index=False, header=False))

    def log_case(self, api_name, mutated_arg, i):
        # 记录错误到日志文件
        record = pd.DataFrame([{
            'timestamp': datetime.now(),
            'api_name': api_name,
            'mutated_arg': str(mutated_arg)

        }])
        '''
        none_bug_case_file = 'none_bug_case/' + api_name + '_' + str(i) + '.txt'
        with open(none_bug_case_file, 'a') as f1:  # 打开文件以追加模式
            # 如果文件是空的，写入头部（列名）
            if f1.tell() == 0:
                f1.write(record.to_csv(sep='\t', index=False))
            else:
                f1.write(record.to_csv(sep='\t', index=False, header=False))
'''
    def _generate_key(self, mutated_arg):
        # 生成参数组合的唯一标识符
        return str(mutated_arg)

def mutate_parameters(args, params_info, mutation_history):
    mutated_args = []
    p_type = ''
    pre_type = ''
    pre_name = ''
    pre_value = 0
    src_type = ''
    reshape_vector = (random.randint(15,50)*2)
    reshape_maxtrix = (reshape_vector, reshape_vector)
    reshape_uvmaxtrix = (reshape_vector//2, reshape_vector//2)
    reshape_2dpoint = (reshape_vector//2, 1, 2)
    reshape_3dpoint = (reshape_vector // 2, 1, 3)
    reshape_img = (reshape_vector, reshape_vector, 3)
    seq_img = (3, reshape_vector, reshape_vector, 3)
    reshape_img_1 = (reshape_vector, reshape_vector, 1)
    reshape_img_2 = (reshape_vector, reshape_vector, 2)
    reshape_mask = (reshape_vector+2, reshape_vector+2)
    reshape_23 = (3, 2)
    reshape_32 = (2, 3)
    for arg, (param_name, param_info) in zip(args, params_info.items()):
        param_type = param_info.get('type')
        strategy = mutation_history.suggest_mutation_strategy(param_type)

        if param_name == 'output':
            continue  # 如果参数名为'output', 则跳过处理
        if ('default' in param_info) and (param_info['default'] == 'None'):
            mutated_param = None

        mutated_param = None
        attempt = 0  # 用于记录尝试次数，防止无限循环
        if 'same' in param_info['type']:
            if (param_name == 'src2') or (param_name == 'img2'):
                p_type = src_type
            else:
                p_type = pre_type
        else:
            p_type = None
        if ('default' in param_info) and (param_info['default'] == 'None'):
            mutated_param = None
        elif param_name == 'lut':
            mutated_param = arg
        elif (pre_name == 'dx') and (param_name == 'dy'):
            mutated_param = 1 - pre_value
        elif ('flag' in param_info) and (param_info['flag'] == 'None'):
            mutated_param = arg
        else:
            if param_info.get('format') == 'numpy.ndarray':
                if (param_name == 'points') or('2D point'in param_info['description']) or ('2D points' in param_info['description']) or ('2xN/Nx2 1-channel' in param_info['description']) or ('corners' in param_name) or ('points (x, y)' in param_info['description']) or ('1xN/Nx1 2-channel' in param_info['description']) :
                    arg = np.resize(arg, reshape_2dpoint)
                elif ('uv_plane' in param_name):
                    arg = np.resize(arg, reshape_uvmaxtrix)
                elif ('3xN/Nx3 1-channel' in param_info['description']) or ('1xN/Nx1 3-channel' in param_info['description']):
                    arg = np.resize(arg, reshape_3dpoint)
                elif '1xNx2 array' in param_info['description']:
                    arg = np.resize(arg, ( 1, reshape_vector, 2))
                elif ('4-element vector' in param_info['description']) :
                    arg = np.resize(arg, reshape_2dpoint)
                elif ('3x4 matrix' in param_info['description']) :
                    arg = np.resize(arg, (4,3))
                elif ('2x4 matrix' in param_info['description']) :
                    arg = np.resize(arg, (4,2))
                elif ('4x4 matrix' in param_info['description']) :
                    arg = np.resize(arg, (4,4))
                elif ('4x3 matrix' in param_info['description']):
                    arg = np.resize(arg, (3, 4))
                elif '2x1 shape' in param_info['description']:
                    arg =  np.full((1,2), reshape_vector).astype(np.int32)
                elif ('warp_matrix 2x3' in param_info['description']):
                    arg = np.eye(2, 3, dtype=np.float32)
                elif ('identity matrix' in param_info['description']) or ('input map' in param_info['description']):
                    arg = np.resize(arg, reshape_maxtrix)
                elif ('3-element vector [x, y, 1]' in param_info['description']):
                    arg = np.resize(arg, reshape_img_1)
                elif ('3x2' in param_info['description']) :
                    arg = np.resize(arg, reshape_32)
                elif ('2x3' in param_info['description']) :
                    arg = np.resize(arg, reshape_23)
                elif ('2x2' in param_info['description']) :
                    arg = np.resize(arg, (2,2))
                elif ('3x1' in param_info['description']) :
                    arg = np.resize(arg, (3,1))
                elif ('matrix' in param_info['description']) or ('3x3' in param_info['description']):
                    arg = np.resize(arg, (3, 3))
                elif ('x3' in param_info['description']):
                    arg = arg
                elif 'Convolution kernel' in param_info['description']:
                    kernelsize = random.choice([3, 5, 7])
                    random_size = [kernelsize,kernelsize]
                    arg = np.resize(arg, random_size)
                elif 'sequence of 8-bit 3-channel'in param_info['description']:
                    arg = np.resize(arg, seq_img)
                elif 'sequence of 8-bit 1-channel' in param_info['description']:
                    arg = (np.random.rand(3,100, 100, 1) * 256).astype(np.uint8)
                elif '8-bit 3-channel'in param_info['description']:
                    arg = np.resize(arg, reshape_img)
                elif '8-bit single-channel'in param_info['description']:
                    arg = np.resize(arg, reshape_img_1)
                elif '2-channel'in param_info['description']:
                    arg = np.resize(arg, reshape_img_2)
                elif '2 pixels taller'in param_info['description']:
                    arg = np.resize(arg, reshape_mask)
                elif ('1D' in param_info['description']) or ('vector' in param_info['description']) or ('array' in param_info['description']) and (len(arg.shape) == 1):
                    arg = np.resize(arg, reshape_vector)
                #elif 'input image'in param_info['description']:
                #    arg = np.resize(arg, reshape_img)
                else:
                    arg = np.resize(arg, reshape_maxtrix)
            while (mutated_param is None) or (not mutation_history.check_combination(mutated_param)) and (attempt < 10):
                if param_info.get('format') == 'numpy.ndarray':
                    if '2x1 shape' in param_info['description']:
                        mutated_param = arg
                    elif ('warp_matrix 2x3' in param_info['description']):
                        mutated_param = np.eye(3, 3, dtype=np.float32)

                    else:
                        mutated_param = apply_ndarray_mutation(arg, param_info, strategy, p_type, param_name).copy()
                    if 'convex polygon' in param_info['description']:
                        mutated_param = mutated_param.reshape(-1, 1, 2).copy()
                    elif ('histogram' in param_info['description']):
                        mutated_param = cv2.calcHist([mutated_param], [0], None, [256], [0, 256])
                elif param_info.get('format') == 'tuple':
                    mutated_param = apply_tuple_mutation(arg, param_info, strategy, param_name, reshape_vector)
                elif 'char' in param_type:
                    mutated_param = apply_char_mutation(arg, param_info, strategy, param_name, reshape_vector)
                elif 'KeyPoint' in param_type:
                    if isinstance(arg, tuple):
                        temp_list = list(arg)
                        random.shuffle(temp_list)
                        mutated_param = tuple(temp_list)
                    else:
                        mutated_param = apply_KeyPoint_type_mutation(arg, param_type, strategy)
                elif 'int' in param_type:
                    mutated_param = apply_int_type_mutation(arg, param_info, param_type, strategy, param_name)
                elif 'float' in param_type:
                    mutated_param = apply_float_type_mutation(arg, param_info, param_type, strategy)
                elif 'bool' in param_type:
                    mutated_param = apply_bool_type_mutation(arg, param_type, strategy)
                elif 'str' in param_type:
                    mutated_param = apply_str_type_mutation(arg, param_type, strategy)
                elif 'double' in param_type:
                    mutated_param = apply_double_type_mutation(arg, param_type, strategy)
                else:
                    mutated_param = arg

                attempt += 1  # 增加尝试次数


        mutated_args.append(mutated_param)
        if param_info.get('format') == 'numpy.ndarray':
            pre_type = mutated_param.dtype.name
        if param_name == 'src' or param_name =='img' or param_name =='src1' or param_name =='img1':
            src_type = mutated_param.dtype.name
        pre_name = param_name
        pre_value = mutated_param

    return mutated_args


'''
def mutate_parameters(args, params_info):
    mutated_args = []

    for arg, (param_name, param_info) in zip(args, params_info.items()):
        # 检查参数类型，并应用相应的突变规则
        if param_info.get('format') == 'numpy.ndarray':
            # 对于 numpy 数组类型的参数
            if arg.dtype == np.uint8:
                # 随机改变 uint8 类型的数组中的一些值
                noise = np.random.randint(-10, 10, arg.shape, dtype=np.int32)
                mutated_arg = np.clip(arg.astype(np.int32) + noise, 0, 255).astype(np.uint8)
            elif arg.dtype in [np.float32, np.float64]:
                # 随机改变 float 类型的数组中的一些值
                noise = np.random.uniform(-1.0, 1.0, arg.shape)
                mutated_arg = arg + noise
            else:
                mutated_arg = arg
        elif param_info.get('format') == 'tuple':
            mutated_arg = tuple(x + random.uniform(-1.0, 1.0) if isinstance(x, (float, int)) else x for x in arg)

        elif param_info.get('type') == 'int':
            # 对于 int 类型的参数
            mutated_arg = arg + random.randint(-10, 10)

        elif param_info.get('type') == 'float':
            # 对于 float 类型的参数
            mutated_arg = arg + random.uniform(-1.0, 1.0)

        elif param_info.get('type') == 'bool':
            # 对于 bool 类型的参数
            mutated_arg = not arg

        elif param_info.get('type') == 'str':
            # 对于 str 类型的参数
            mutated_arg = ''.join(random.choices(string.ascii_letters + string.digits, k=len(arg)))

        elif param_info.get('type') == 'KeyPoint':
            # 对于 cv2.KeyPoint 类型的参数
            mutated_arg = cv2.KeyPoint(x=arg.pt[0] + random.uniform(-5, 5),
                                       y=arg.pt[1] + random.uniform(-5, 5),
                                       size=arg.size + random.uniform(-1, 1))

        else:
            # 如果参数类型不适用于以上任何规则，保持原样
            mutated_arg = arg

        mutated_args.append(mutated_arg)

    return mutated_args

'''